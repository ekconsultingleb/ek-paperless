import streamlit as st
import pandas as pd
import numpy as np
from supabase import Client
from io import BytesIO
import openpyxl
from dpos_simulation import tab_final_report

# ─────────────────────────────────────────────
#  CONSTANTS
# ─────────────────────────────────────────────
EK_DARK = "#1B252C"
EK_SAND = "#E3C5AD"

TABS = ["Unit Costs", "Sub Recipes", "Recipes", "Sessions", "Final Report"]

# ─────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────

def _client_id(st_session) -> int:
    """Pull client_id from Paperless session state."""
    return st_session.get("selected_client_id") or st_session.get("client_id")


def _require_client(client_id):
    if not client_id:
        st.warning("Please select a client first.")
        st.stop()


def _sb(supabase: Client):
    """Shorthand — just pass through."""
    return supabase


def fmt_cost(val):
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return "—"
    return f"{val:,.4f}"


def fmt_num(val, decimals=2):
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return "—"
    return f"{val:,.{decimals}f}"


def yield_badge(yield_pct):
    if yield_pct is None:
        return ""
    color = "#d4f0d4" if yield_pct >= 90 else "#fff3cd" if yield_pct >= 70 else "#f8d7da"
    text_color = "#1a5c1a" if yield_pct >= 90 else "#856404" if yield_pct >= 70 else "#842029"
    return f'<span style="background:{color};color:{text_color};padding:1px 6px;border-radius:4px;font-size:11px;font-weight:500">{yield_pct:.1f}%</span>'


# ─────────────────────────────────────────────
#  DATA LOADERS
# ─────────────────────────────────────────────

@st.cache_data(ttl=60, show_spinner=False)
def load_unit_costs(_supabase: Client, client_id: int) -> pd.DataFrame:
    res = _supabase.table("dpos_unit_costs") \
        .select("*") \
        .eq("client_id", client_id) \
        .order("category") \
        .execute()
    if not res.data:
        return pd.DataFrame()
    return pd.DataFrame(res.data)


@st.cache_data(ttl=60, show_spinner=False)
def load_sub_recipes(_supabase: Client, client_id: int) -> pd.DataFrame:
    res = _supabase.table("dpos_sub_recipes") \
        .select("*") \
        .eq("client_id", client_id) \
        .order("product_name") \
        .execute()
    if not res.data:
        return pd.DataFrame()
    return pd.DataFrame(res.data)


@st.cache_data(ttl=60, show_spinner=False)
def load_recipes(_supabase: Client, client_id: int) -> pd.DataFrame:
    res = _supabase.table("dpos_recipes") \
        .select("*") \
        .eq("client_id", client_id) \
        .order("category") \
        .execute()
    if not res.data:
        return pd.DataFrame()
    return pd.DataFrame(res.data)


@st.cache_data(ttl=60, show_spinner=False)
def load_sessions(_supabase: Client, client_id: int) -> pd.DataFrame:
    res = _supabase.table("dpos_sessions") \
        .select("*") \
        .eq("client_id", client_id) \
        .order("created_at", desc=True) \
        .execute()
    if not res.data:
        return pd.DataFrame()
    return pd.DataFrame(res.data)


@st.cache_data(ttl=120, show_spinner=False)
def load_clients(_supabase: Client) -> pd.DataFrame:
    res = _supabase.table("clients").select("id, name").order("name").execute()
    if not res.data:
        return pd.DataFrame()
    return pd.DataFrame(res.data)


# ─────────────────────────────────────────────
#  EXCEL IMPORTER — UNIT COSTS
# ─────────────────────────────────────────────

def parse_unit_cost_excel(file_bytes: bytes) -> pd.DataFrame:
    """
    Parse the Unit Cost sheet from a D-POS Excel file.
    Returns a clean DataFrame ready for upsert.
    Expected columns: Category, Group, Product Description,
    Qty Inv, Unit, Qty buy, Avg Cost, Rate
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, keep_vba=False)

    sheet_name = None
    for name in wb.sheetnames:
        if "unit" in name.lower() and "cost" in name.lower():
            sheet_name = name
            break
    if not sheet_name:
        raise ValueError("Could not find 'Unit Cost' sheet in the uploaded file.")

    ws = wb[sheet_name]
    rows = []
    header_found = False
    col_map = {}

    for row in ws.iter_rows(values_only=True):
        # Skip fully empty rows
        if all(v is None for v in row):
            continue

        # Find header row
        if not header_found:
            row_lower = [str(c).lower().strip() if c is not None else "" for c in row]
            if "category" in row_lower and "product description" in row_lower:
                header_found = True
                for i, h in enumerate(row_lower):
                    if "category" in h:
                        col_map["category"] = i
                    elif "group" in h:
                        col_map["group_name"] = i
                    elif "product description" in h:
                        col_map["product_description"] = i
                    elif "qty inv" in h:
                        col_map["qty_inv"] = i
                    elif h == "unit":
                        col_map["unit"] = i
                    elif "qty buy" in h:
                        col_map["qty_buy"] = i
                    elif "avg cost" in h:
                        col_map["avg_cost_lbp"] = i
                    elif "rate" in h:
                        col_map["rate"] = i
            continue

        # Data rows
        def get(key):
            idx = col_map.get(key)
            if idx is None:
                return None
            val = row[idx]
            if isinstance(val, str) and val.strip() == "":
                return None
            return val

        desc = get("product_description")
        if desc is None:
            continue
        # Skip formula strings
        if isinstance(desc, str) and desc.startswith("="):
            continue

        rows.append({
            "category": get("category"),
            "group_name": get("group_name"),
            "product_description": str(desc).strip(),
            "qty_inv": get("qty_inv"),
            "unit": get("unit"),
            "qty_buy": get("qty_buy"),
            "avg_cost_lbp": get("avg_cost_lbp"),
            "rate": get("rate") or 90000,
        })

    wb.close()

    if not rows:
        raise ValueError("No data rows found in Unit Cost sheet.")

    df = pd.DataFrame(rows)
    # Coerce numeric columns
    for col in ["qty_inv", "qty_buy", "avg_cost_lbp", "rate"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    # Default rate
    df["rate"] = df["rate"].fillna(90000)
    return df


def upsert_unit_costs(supabase: Client, client_id: int, df: pd.DataFrame) -> tuple[int, int]:
    """
    Delete existing unit costs for client, insert fresh batch.
    Returns (inserted, skipped).
    """
    # Delete existing
    supabase.table("dpos_unit_costs") \
        .delete() \
        .eq("client_id", client_id) \
        .execute()

    records = []
    skipped = 0
    for _, row in df.iterrows():
        if pd.isna(row.get("product_description")) or not row["product_description"]:
            skipped += 1
            continue
        rec = {
            "client_id": client_id,
            "category": row.get("category"),
            "group_name": row.get("group_name"),
            "product_description": row["product_description"],
            "qty_inv": None if pd.isna(row.get("qty_inv", float("nan"))) else float(row["qty_inv"]),
            "unit": row.get("unit"),
            "qty_buy": None if pd.isna(row.get("qty_buy", float("nan"))) else float(row["qty_buy"]),
            "avg_cost_lbp": None if pd.isna(row.get("avg_cost_lbp", float("nan"))) else float(row["avg_cost_lbp"]),
            "rate": float(row.get("rate") or 90000),
            "currency": "LBP",
            "show_in_report": True,
        }
        records.append(rec)

    if records:
        # Batch in chunks of 500
        for i in range(0, len(records), 500):
            supabase.table("dpos_unit_costs").insert(records[i:i+500]).execute()

    return len(records), skipped


# ─────────────────────────────────────────────
#  TAB 1 — UNIT COSTS
# ─────────────────────────────────────────────

def tab_unit_costs(supabase: Client, client_id: int):
    st.markdown("#### Unit Costs")

    col_import, col_spacer = st.columns([2, 5])

    with col_import:
        with st.expander("Import from Excel", expanded=False):
            uploaded = st.file_uploader(
                "Upload D-POS Excel file (.xlsm / .xlsx)",
                type=["xlsm", "xlsx"],
                key="uc_excel_upload",
            )
            if uploaded:
                st.caption(f"File: {uploaded.name} ({uploaded.size / 1024:.1f} KB)")
                if st.button("Parse & Preview", key="uc_parse_btn"):
                    with st.spinner("Reading file…"):
                        try:
                            df_preview = parse_unit_cost_excel(uploaded.read())
                            st.session_state["uc_preview_df"] = df_preview
                            st.session_state["uc_preview_file"] = uploaded.name
                        except Exception as e:
                            st.error(f"Parse error: {e}")

            if "uc_preview_df" in st.session_state:
                df_prev = st.session_state["uc_preview_df"]
                st.success(f"Preview: {len(df_prev)} rows from **{st.session_state.get('uc_preview_file', '')}**")
                st.dataframe(
                    df_prev[["category", "product_description", "qty_inv", "unit", "avg_cost_lbp", "rate"]].head(10),
                    use_container_width=True,
                    hide_index=True,
                )
                st.caption(f"Showing first 10 of {len(df_prev)} rows.")

                if st.button("Push to Supabase", type="primary", key="uc_push_btn"):
                    with st.spinner("Uploading…"):
                        try:
                            inserted, skipped = upsert_unit_costs(supabase, client_id, df_prev)
                            st.success(f"Done — {inserted} rows inserted, {skipped} skipped.")
                            del st.session_state["uc_preview_df"]
                            load_unit_costs.clear()
                            st.rerun()
                        except Exception as e:
                            st.error(f"Upload failed: {e}")

    # ── Load & display ──
    df = load_unit_costs(supabase, client_id)

    if df.empty:
        st.info("No unit costs found for this client. Import from Excel above or add manually.")
        _manual_add_unit_cost(supabase, client_id)
        return

    # ── Filters ──
    f1, f2, f3 = st.columns([2, 2, 3])
    with f1:
        cats = ["All"] + sorted(df["category"].dropna().unique().tolist())
        cat_filter = st.selectbox("Category", cats, key="uc_cat_filter")
    with f2:
        groups = ["All"] + sorted(df["group_name"].dropna().unique().tolist())
        grp_filter = st.selectbox("Group", groups, key="uc_grp_filter")
    with f3:
        search = st.text_input("Search ingredient", placeholder="e.g. chicken, mayo…", key="uc_search")

    filtered = df.copy()
    if cat_filter != "All":
        filtered = filtered[filtered["category"] == cat_filter]
    if grp_filter != "All":
        filtered = filtered[filtered["group_name"] == grp_filter]
    if search:
        filtered = filtered[filtered["product_description"].str.contains(search, case=False, na=False)]

    st.caption(f"Showing {len(filtered)} of {len(df)} items")

    # ── Display table ──
    display_cols = ["category", "group_name", "product_description", "qty_inv", "unit", "qty_buy", "avg_cost_lbp", "rate", "unit_cost_usd", "usage_cost_usd"]
    display_cols = [c for c in display_cols if c in filtered.columns]

    rename_map = {
        "category": "Category",
        "group_name": "Group",
        "product_description": "Description",
        "qty_inv": "Qty Inv",
        "unit": "Unit",
        "qty_buy": "Qty Buy",
        "avg_cost_lbp": "Avg Cost",
        "rate": "Rate",
        "unit_cost_usd": "Unit Cost $",
        "usage_cost_usd": "Usage Cost $",
    }

    show_df = filtered[display_cols].rename(columns=rename_map)
    st.dataframe(show_df, use_container_width=True, hide_index=True)

    # ── Inline edit / delete ──
    with st.expander("Edit or delete an item", expanded=False):
        items = filtered["product_description"].tolist()
        if not items:
            st.caption("No items match current filter.")
        else:
            selected_desc = st.selectbox("Select item", items, key="uc_edit_select")
            row = filtered[filtered["product_description"] == selected_desc].iloc[0]
            _edit_unit_cost_form(supabase, client_id, row)

    # ── Manual add ──
    with st.expander("Add new item manually", expanded=False):
        _manual_add_unit_cost(supabase, client_id)


def _edit_unit_cost_form(supabase: Client, client_id: int, row: pd.Series):
    with st.form("uc_edit_form"):
        c1, c2, c3 = st.columns(3)
        with c1:
            cat = st.text_input("Category", value=row.get("category") or "", key="uc_e_cat")
            grp = st.text_input("Group", value=row.get("group_name") or "", key="uc_e_grp")
        with c2:
            desc = st.text_input("Description", value=row.get("product_description") or "", key="uc_e_desc")
            unit = st.text_input("Unit", value=row.get("unit") or "", key="uc_e_unit")
        with c3:
            avg_cost = st.number_input("Avg Cost (LBP)", value=float(row.get("avg_cost_lbp") or 0), key="uc_e_cost")
            rate = st.number_input("Rate", value=float(row.get("rate") or 90000), key="uc_e_rate")

        col_save, col_del = st.columns(2)
        with col_save:
            save = st.form_submit_button("Save changes", type="primary")
        with col_del:
            delete = st.form_submit_button("Delete item", type="secondary")

    if save:
        supabase.table("dpos_unit_costs").update({
            "category": cat or None,
            "group_name": grp or None,
            "product_description": desc,
            "unit": unit or None,
            "avg_cost_lbp": avg_cost,
            "rate": rate,
        }).eq("id", int(row["id"])).execute()
        st.success("Saved.")
        load_unit_costs.clear()
        st.rerun()

    if delete:
        supabase.table("dpos_unit_costs").delete().eq("id", int(row["id"])).execute()
        st.success("Deleted.")
        load_unit_costs.clear()
        st.rerun()


def _manual_add_unit_cost(supabase: Client, client_id: int):
    with st.form("uc_add_form", clear_on_submit=True):
        st.markdown("**New item**")
        c1, c2, c3 = st.columns(3)
        with c1:
            cat = st.text_input("Category", key="uc_a_cat")
            grp = st.text_input("Group", key="uc_a_grp")
        with c2:
            desc = st.text_input("Description *", key="uc_a_desc")
            unit = st.text_input("Unit (g / ml / Btl…)", key="uc_a_unit")
        with c3:
            avg_cost = st.number_input("Avg Cost (LBP)", min_value=0.0, key="uc_a_cost")
            rate = st.number_input("Rate", value=90000.0, key="uc_a_rate")
            qty_inv = st.number_input("Qty Inv", min_value=0.0, key="uc_a_qty")

        submitted = st.form_submit_button("Add item", type="primary")

    if submitted:
        if not desc.strip():
            st.error("Description is required.")
            return
        supabase.table("dpos_unit_costs").insert({
            "client_id": client_id,
            "category": cat or None,
            "group_name": grp or None,
            "product_description": desc.strip(),
            "unit": unit or None,
            "qty_inv": qty_inv or None,
            "avg_cost_lbp": avg_cost,
            "rate": rate,
            "currency": "LBP",
            "show_in_report": True,
        }).execute()
        st.success(f"Added: {desc}")
        load_unit_costs.clear()
        st.rerun()


# ─────────────────────────────────────────────
#  TAB 2 — SUB RECIPES
# ─────────────────────────────────────────────

def tab_sub_recipes(supabase: Client, client_id: int):
    st.markdown("#### Sub Recipes")
    st.caption("Intermediate productions — batch-costed items used as ingredients in recipes.")

    df = load_sub_recipes(supabase, client_id)
    uc_df = load_unit_costs(supabase, client_id)

    # Build unit cost lookup
    uc_lookup = {}
    if not uc_df.empty:
        uc_lookup = dict(zip(
            uc_df["product_description"].str.lower().str.strip(),
            uc_df["usage_cost_usd"].fillna(uc_df["avg_cost_lbp"])
        ))

    if df.empty:
        st.info("No sub recipes found. Add one below.")
    else:
        # Group by product_name
        products = sorted(df["product_name"].dropna().unique().tolist())

        search = st.text_input("Search sub recipe", placeholder="e.g. lemon juice, bechamel…", key="sr_search")
        if search:
            products = [p for p in products if search.lower() in p.lower()]

        for prod in products:
            lines = df[df["product_name"] == prod]
            total_batch = lines["batch_cost"].sum() if "batch_cost" in lines.columns else None
            prep_qty = lines.iloc[0].get("prepared_qty") if len(lines) > 0 else None
            cost_for_1 = lines.iloc[0].get("cost_for_1") if len(lines) > 0 else None

            with st.expander(f"**{prod}**  —  batch cost: {fmt_num(total_batch)}  |  cost/1: {fmt_num(cost_for_1, 4)}", expanded=False):
                display = lines[["ingredient_description", "net_w", "gross_w", "yield_pct", "unit_name", "avg_cost", "batch_cost"]].copy()
                display.rename(columns={
                    "ingredient_description": "Ingredient",
                    "net_w": "Net W",
                    "gross_w": "Gross W",
                    "yield_pct": "Yield %",
                    "unit_name": "Unit",
                    "avg_cost": "Avg Cost",
                    "batch_cost": "Batch Cost",
                }, inplace=True)
                st.dataframe(display, use_container_width=True, hide_index=True)

                col_info, col_del = st.columns([3, 1])
                with col_info:
                    if prep_qty:
                        st.caption(f"Prepared qty: {fmt_num(prep_qty)} {lines.iloc[0].get('prepared_unit', '')}")
                with col_del:
                    if st.button("Delete sub recipe", key=f"sr_del_{prod}", type="secondary"):
                        ids = lines["id"].tolist()
                        for rid in ids:
                            supabase.table("dpos_sub_recipes").delete().eq("id", rid).execute()
                        load_sub_recipes.clear()
                        st.rerun()

    # ── Add new sub recipe ──
    with st.expander("Add / append sub recipe lines", expanded=False):
        _add_sub_recipe_form(supabase, client_id, uc_lookup)


def _add_sub_recipe_form(supabase: Client, client_id: int, uc_lookup: dict):
    st.markdown("**Sub recipe header**")
    c1, c2, c3 = st.columns(3)
    with c1:
        prod_name = st.text_input("Production name *", key="sr_prod_name")
    with c2:
        prep_qty = st.number_input("Prepared qty", min_value=0.0, key="sr_prep_qty")
        prep_unit = st.text_input("Prepared unit (Ltr / kg…)", key="sr_prep_unit")

    st.markdown("**Ingredient lines**")

    if "sr_lines" not in st.session_state:
        st.session_state["sr_lines"] = [{}]

    lines_out = []
    for i, line in enumerate(st.session_state["sr_lines"]):
        lc1, lc2, lc3, lc4, lc5 = st.columns([3, 1, 1, 1, 1])
        with lc1:
            ing = st.text_input("Ingredient", value=line.get("ing", ""), key=f"sr_ing_{i}")
        with lc2:
            net_w = st.number_input("Net W", value=float(line.get("net_w", 0)), min_value=0.0, key=f"sr_nw_{i}")
        with lc3:
            gross_w = st.number_input("Gross W", value=float(line.get("gross_w", 0)), min_value=0.0, key=f"sr_gw_{i}")
        with lc4:
            unit = st.text_input("Unit", value=line.get("unit", ""), key=f"sr_unit_{i}")
        with lc5:
            # Auto-lookup cost
            auto_cost = uc_lookup.get(ing.lower().strip(), 0.0) if ing else 0.0
            avg_cost = st.number_input("Avg Cost", value=float(line.get("avg_cost", auto_cost or 0)), key=f"sr_cost_{i}")

        lines_out.append({"ing": ing, "net_w": net_w, "gross_w": gross_w, "unit": unit, "avg_cost": avg_cost})

    col_add, col_save = st.columns(2)
    with col_add:
        if st.button("+ Add line", key="sr_add_line"):
            st.session_state["sr_lines"].append({})
            st.rerun()
    with col_save:
        if st.button("Save sub recipe", type="primary", key="sr_save"):
            if not prod_name.strip():
                st.error("Production name is required.")
                return
            records = []
            for ln in lines_out:
                if not ln["ing"].strip():
                    continue
                gw = ln["gross_w"] or ln["net_w"] or 0
                nw = ln["net_w"] or gw
                records.append({
                    "client_id": client_id,
                    "product_name": prod_name.strip(),
                    "ingredient_description": ln["ing"].strip(),
                    "net_w": nw or None,
                    "gross_w": gw or None,
                    "unit_name": ln["unit"] or None,
                    "avg_cost": ln["avg_cost"] or None,
                    "prepared_qty": prep_qty or None,
                    "prepared_unit": prep_unit or None,
                })
            if records:
                supabase.table("dpos_sub_recipes").insert(records).execute()
                st.success(f"Saved {len(records)} lines for '{prod_name}'.")
                st.session_state["sr_lines"] = [{}]
                load_sub_recipes.clear()
                st.rerun()
            else:
                st.error("No valid ingredient lines to save.")


# ─────────────────────────────────────────────
#  TAB 3 — RECIPES
# ─────────────────────────────────────────────

def tab_recipes(supabase: Client, client_id: int):
    st.markdown("#### Recipes")

    df = load_recipes(supabase, client_id)
    uc_df = load_unit_costs(supabase, client_id)
    sr_df = load_sub_recipes(supabase, client_id)

    # Build lookups
    uc_lookup = {}
    if not uc_df.empty:
        uc_lookup = dict(zip(
            uc_df["product_description"].str.lower().str.strip(),
            uc_df["usage_cost_usd"].fillna(uc_df["avg_cost_lbp"])
        ))

    sr_lookup = {}
    if not sr_df.empty:
        for prod, grp in sr_df.groupby("product_name"):
            c1_val = grp.iloc[0].get("cost_for_1")
            if c1_val:
                sr_lookup[prod.lower().strip()] = float(c1_val)

    if df.empty:
        st.info("No recipes found. Add one below.")
        _add_recipe_form(supabase, client_id, uc_lookup, sr_lookup, uc_df, sr_df)
        return

    # ── Filters ──
    f1, f2, f3 = st.columns([2, 2, 3])
    with f1:
        cats = ["All"] + sorted(df["category"].dropna().unique().tolist())
        cat_filter = st.selectbox("Category", cats, key="rec_cat_filter")
    with f2:
        groups = ["All"] + sorted(df["group_name"].dropna().unique().tolist())
        grp_filter = st.selectbox("Group", groups, key="rec_grp_filter")
    with f3:
        search = st.text_input("Search menu item", placeholder="e.g. burger, salad…", key="rec_search")

    filtered = df.copy()
    if cat_filter != "All":
        filtered = filtered[filtered["category"] == cat_filter]
    if grp_filter != "All":
        filtered = filtered[filtered["group_name"] == grp_filter]
    if search:
        filtered = filtered[filtered["menu_item"].str.contains(search, case=False, na=False)]

    menu_items = sorted(filtered["menu_item"].dropna().unique().tolist())
    st.caption(f"{len(menu_items)} menu items · {len(filtered)} ingredient lines")

    for item in menu_items:
        lines = filtered[filtered["menu_item"] == item]
        total_cost = lines["total_cost"].sum() if "total_cost" in lines.columns else None
        cat_label = lines.iloc[0].get("category", "") or ""
        grp_label = lines.iloc[0].get("group_name", "") or ""

        header = f"**{item}**"
        if cat_label or grp_label:
            header += f"  <span style='color:var(--text-color);opacity:0.5;font-size:12px'>{cat_label} / {grp_label}</span>"
        cost_label = f"  —  recipe cost: **{fmt_num(total_cost)}**" if total_cost else ""

        with st.expander(f"{item}  —  cost: {fmt_num(total_cost)}", expanded=False):
            # Build display rows with yield badge
            rows_display = []
            for _, ln in lines.iterrows():
                yield_pct = ln.get("yield_pct")
                rows_display.append({
                    "Ingredient": ln.get("ingredient_description", ""),
                    "Net W": fmt_num(ln.get("net_w"), 2),
                    "Gross W": fmt_num(ln.get("gross_w"), 2),
                    "Yield %": f"{yield_pct:.1f}%" if yield_pct else "—",
                    "Unit": ln.get("unit", ""),
                    "Avg Cost": fmt_cost(ln.get("avg_cost")),
                    "Total Cost": fmt_num(ln.get("total_cost"), 4),
                })

            st.dataframe(pd.DataFrame(rows_display), use_container_width=True, hide_index=True)

            # Yield warnings
            has_yield = lines["yield_pct"].notna().any() if "yield_pct" in lines.columns else False
            if has_yield:
                low_yield = lines[lines["yield_pct"].notna() & (lines["yield_pct"] < 70)]
                if not low_yield.empty:
                    for _, lrow in low_yield.iterrows():
                        st.warning(f"Low yield on {lrow['ingredient_description']}: {lrow['yield_pct']:.1f}% — check trim loss.")

            col_del, _ = st.columns([1, 3])
            with col_del:
                if st.button("Delete recipe", key=f"rec_del_{item}", type="secondary"):
                    ids = lines["id"].tolist()
                    for rid in ids:
                        supabase.table("dpos_recipes").delete().eq("id", rid).execute()
                    load_recipes.clear()
                    st.rerun()

    st.divider()
    with st.expander("Add new recipe", expanded=False):
        _add_recipe_form(supabase, client_id, uc_lookup, sr_lookup, uc_df, sr_df)


def _add_recipe_form(supabase: Client, client_id: int, uc_lookup: dict, sr_lookup: dict, uc_df: pd.DataFrame, sr_df: pd.DataFrame):
    st.markdown("**Recipe header**")
    h1, h2, h3 = st.columns(3)
    with h1:
        menu_item = st.text_input("Menu item name *", key="rec_a_item")
        category = st.text_input("Category", key="rec_a_cat")
    with h2:
        group_name = st.text_input("Group", key="rec_a_grp")

    st.markdown("**Ingredient lines**")
    st.caption("Ingredient can be a raw material (from Unit Costs) or a sub recipe.")

    # Ingredient options for autocomplete
    uc_options = uc_df["product_description"].tolist() if not uc_df.empty else []
    sr_options = sr_df["product_name"].dropna().unique().tolist() if not sr_df.empty else []
    all_options = sorted(set(uc_options + sr_options))

    if "rec_lines" not in st.session_state:
        st.session_state["rec_lines"] = [{}]

    lines_out = []
    for i, line in enumerate(st.session_state["rec_lines"]):
        lc1, lc2, lc3, lc4, lc5 = st.columns([3, 1, 1, 1, 1])
        with lc1:
            if all_options:
                ing = st.selectbox(
                    "Ingredient",
                    [""] + all_options,
                    index=0,
                    key=f"rec_ing_{i}",
                )
            else:
                ing = st.text_input("Ingredient", key=f"rec_ing_{i}")
        with lc2:
            net_w = st.number_input("Net W", value=0.0, min_value=0.0, key=f"rec_nw_{i}")
        with lc3:
            gross_w = st.number_input("Gross W", value=0.0, min_value=0.0, key=f"rec_gw_{i}")
        with lc4:
            unit = st.text_input("Unit", value="", key=f"rec_unit_{i}")
        with lc5:
            # Auto-lookup: check unit costs first, then sub recipes
            auto_cost = 0.0
            if ing:
                key_lower = ing.lower().strip()
                auto_cost = uc_lookup.get(key_lower) or sr_lookup.get(key_lower) or 0.0
            avg_cost = st.number_input("Avg Cost", value=float(auto_cost), key=f"rec_cost_{i}")

        # Show yield hint inline
        if net_w > 0 and gross_w > 0 and net_w != gross_w:
            yp = (net_w / gross_w) * 100
            color = "green" if yp >= 90 else "orange" if yp >= 70 else "red"
            st.markdown(f"<small style='color:{color}'>yield: {yp:.1f}%</small>", unsafe_allow_html=True)

        lines_out.append({"ing": ing, "net_w": net_w, "gross_w": gross_w, "unit": unit, "avg_cost": avg_cost})

    col_add, col_save = st.columns(2)
    with col_add:
        if st.button("+ Add ingredient line", key="rec_add_line"):
            st.session_state["rec_lines"].append({})
            st.rerun()
    with col_save:
        if st.button("Save recipe", type="primary", key="rec_save"):
            if not menu_item.strip():
                st.error("Menu item name is required.")
                return
            records = []
            for ln in lines_out:
                if not ln["ing"] or not str(ln["ing"]).strip():
                    continue
                gw = ln["gross_w"] or ln["net_w"] or 0
                nw = ln["net_w"] or gw
                records.append({
                    "client_id": client_id,
                    "category": category or None,
                    "group_name": group_name or None,
                    "menu_item": menu_item.strip(),
                    "ingredient_description": str(ln["ing"]).strip(),
                    "net_w": nw or None,
                    "gross_w": gw or None,
                    "unit": ln["unit"] or None,
                    "avg_cost": ln["avg_cost"] or None,
                })
            if records:
                supabase.table("dpos_recipes").insert(records).execute()
                st.success(f"Saved recipe '{menu_item}' with {len(records)} ingredients.")
                st.session_state["rec_lines"] = [{}]
                load_recipes.clear()
                st.rerun()
            else:
                st.error("No valid ingredient lines.")


# ─────────────────────────────────────────────
#  TAB 4 — SESSIONS
# ─────────────────────────────────────────────

def tab_sessions(supabase: Client, client_id: int):
    st.markdown("#### Pricing Sessions")
    st.caption("Each session is a repricing exercise. Create one to start simulating.")

    df = load_sessions(supabase, client_id)

    # ── Create new session ──
    with st.expander("Create new session", expanded=df.empty):
        with st.form("sess_create_form", clear_on_submit=True):
            s1, s2, s3 = st.columns(3)
            with s1:
                sess_name = st.text_input("Session name *", placeholder="e.g. April 2026 Repricing")
            with s2:
                vat_rate = st.number_input("VAT rate %", min_value=0.0, max_value=30.0, value=11.0, step=0.5)
            with s3:
                target_pct = st.number_input("Target food cost %", min_value=5.0, max_value=80.0, value=30.0, step=1.0)
            notes = st.text_area("Notes (optional)", height=60)
            submitted = st.form_submit_button("Create session", type="primary")

        if submitted:
            if not sess_name.strip():
                st.error("Session name is required.")
            else:
                supabase.table("dpos_sessions").insert({
                    "client_id": client_id,
                    "session_name": sess_name.strip(),
                    "vat_rate": vat_rate / 100,
                    "target_cost_pct": target_pct / 100,
                    "status": "draft",
                    "notes": notes or None,
                    "created_by": st.session_state.get("username", "EK Team"),
                }).execute()
                st.success(f"Session '{sess_name}' created.")
                load_sessions.clear()
                st.rerun()

    if df.empty:
        st.info("No sessions yet.")
        return

    # ── Session list ──
    status_colors = {"draft": "#fff3cd", "approved": "#d4edda", "archived": "#e2e3e5"}
    status_text = {"draft": "#856404", "approved": "#155724", "archived": "#383d41"}

    for _, sess in df.iterrows():
        status = sess.get("status", "draft")
        bg = status_colors.get(status, "#fff")
        tc = status_text.get(status, "#000")

        with st.expander(
            f"**{sess['session_name']}**  —  {sess.get('created_at', '')[:10]}  |  status: {status}",
            expanded=False
        ):
            sc1, sc2, sc3 = st.columns(3)
            with sc1:
                st.metric("VAT rate", f"{float(sess.get('vat_rate', 0)) * 100:.1f}%")
            with sc2:
                st.metric("Target cost %", f"{float(sess.get('target_cost_pct', 0.3)) * 100:.1f}%")
            with sc3:
                st.metric("Status", status.upper())

            if sess.get("notes"):
                st.caption(f"Notes: {sess['notes']}")

            # Status actions
            act1, act2, act3 = st.columns(3)
            with act1:
                if status == "draft" and st.button("Mark approved", key=f"sess_approve_{sess['id']}"):
                    supabase.table("dpos_sessions").update({"status": "approved"}).eq("id", sess["id"]).execute()
                    load_sessions.clear()
                    st.rerun()
            with act2:
                if status != "archived" and st.button("Archive", key=f"sess_archive_{sess['id']}"):
                    supabase.table("dpos_sessions").update({"status": "archived"}).eq("id", sess["id"]).execute()
                    load_sessions.clear()
                    st.rerun()
            with act3:
                if st.button("Delete session", key=f"sess_del_{sess['id']}", type="secondary"):
                    supabase.table("dpos_sessions").delete().eq("id", sess["id"]).execute()
                    load_sessions.clear()
                    st.rerun()

            # Show cost overrides if any
            overrides_res = supabase.table("dpos_cost_overrides") \
                .select("*").eq("session_id", sess["id"]).execute()
            if overrides_res.data:
                st.markdown("**Cost overrides (what-if)**")
                ov_df = pd.DataFrame(overrides_res.data)
                ov_df["change"] = ((ov_df["predicted_cost"] - ov_df["original_cost"]) / ov_df["original_cost"] * 100).round(1)
                st.dataframe(
                    ov_df[["product_description", "original_cost", "predicted_cost", "change", "notes"]].rename(columns={
                        "product_description": "Ingredient",
                        "original_cost": "Current Cost",
                        "predicted_cost": "Predicted Cost",
                        "change": "Change %",
                        "notes": "Notes",
                    }),
                    use_container_width=True,
                    hide_index=True,
                )

            # Add override
            with st.expander("Add cost override", expanded=False):
                _add_override_form(supabase, int(sess["id"]), client_id)


def _add_override_form(supabase: Client, session_id: int, client_id: int):
    uc_df = load_unit_costs(supabase, client_id)
    uc_options = uc_df["product_description"].tolist() if not uc_df.empty else []

    with st.form(f"override_form_{session_id}", clear_on_submit=True):
        o1, o2, o3 = st.columns(3)
        with o1:
            if uc_options:
                ingredient = st.selectbox("Ingredient", [""] + uc_options, key=f"ov_ing_{session_id}")
            else:
                ingredient = st.text_input("Ingredient", key=f"ov_ing_{session_id}")
        with o2:
            original = st.number_input("Current cost", min_value=0.0, key=f"ov_orig_{session_id}")
            predicted = st.number_input("Predicted cost", min_value=0.0, key=f"ov_pred_{session_id}")
        with o3:
            notes = st.text_input("Notes", key=f"ov_notes_{session_id}")

        submitted = st.form_submit_button("Add override", type="primary")

    if submitted:
        if not ingredient or not str(ingredient).strip():
            st.error("Ingredient required.")
            return
        supabase.table("dpos_cost_overrides").insert({
            "session_id": session_id,
            "product_description": str(ingredient).strip(),
            "original_cost": original or None,
            "predicted_cost": predicted,
            "notes": notes or None,
        }).execute()
        st.success("Override added.")
        st.rerun()


# ─────────────────────────────────────────────
#  MAIN ENTRY POINT
# ─────────────────────────────────────────────

def show_dpos(supabase: Client):
    """
    Main entry point — called from Paperless app.py navigation.
    Expects st.session_state to contain:
      - selected_client_id (or client_id)
      - user_role
    """

    # ── Role gate ──
    user_role = st.session_state.get("user_role", "")
    if user_role not in ("admin", "ek_team", "superadmin"):
        st.error("Access restricted to EK team members.")
        st.stop()

    # ── Header ──
    st.markdown(
        f"""
        <div style="background:{EK_DARK};padding:16px 20px;border-radius:10px;margin-bottom:20px">
          <span style="color:{EK_SAND};font-size:20px;font-weight:600;letter-spacing:0.5px">
            D-POS Pricing Studio
          </span>
          <span style="color:{EK_SAND};opacity:0.6;font-size:13px;margin-left:12px">
            Dynamic Price Optimization Solution
          </span>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # ── Client selector ──
    clients_df = load_clients(supabase)

    if clients_df.empty:
        st.error("No clients found.")
        st.stop()

    client_options = dict(zip(clients_df["name"], clients_df["id"]))

    # Pre-select from session state if available
    default_name = None
    existing_id = _client_id(st.session_state)
    if existing_id:
        match = clients_df[clients_df["id"] == existing_id]
        if not match.empty:
            default_name = match.iloc[0]["name"]

    selected_name = st.selectbox(
        "Client",
        list(client_options.keys()),
        index=list(client_options.keys()).index(default_name) if default_name in client_options else 0,
        key="dpos_client_selector",
    )
    client_id = client_options[selected_name]

    # Store for downstream use
    st.session_state["dpos_client_id"] = client_id

    st.divider()

    # ── Tabs ──
    tab1, tab2, tab3, tab4, tab5 = st.tabs(TABS)

    with tab1:
        tab_unit_costs(supabase, client_id)

    with tab2:
        tab_sub_recipes(supabase, client_id)

    with tab3:
        tab_recipes(supabase, client_id)

    with tab4:
        tab_sessions(supabase, client_id)

    with tab5:
        tab_final_report(supabase, client_id)
